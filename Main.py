import os
import time
import random
import datetime
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver

chrome_options = webdriver.ChromeOptions()
# 开启无头模式
chrome_options.add_argument("--headless")
# 不加载图片,提高效率
prefs = {"profile.managed_default_content_settings.images": 2}
chrome_options.add_experimental_option("prefs", prefs)
chrome_options.add_argument(
    'User-Agent:"Mozilla/5.0(Windows;U;WindowsNT6.1;en-us)AppleWebKit/534.50(KHTML,likeGecko)Version/5.1Safari/534.50"')
browser = webdriver.Chrome(executable_path=r'chromedriver.exe',
                           chrome_options=chrome_options)

file_name = 'data.xlsx'  # 存储数据文件名
today = datetime.date.today()  # 启动date

real_max_page_num = 100  # 实际最大页码数
row = 2  # 表格数据开始行

base_url = 'https://www.971cf.com'  # 网站总换地址
mid = '/htm'  # 中间分隔
suffix = '.htm'  # 后缀
# 类别字典
type_dict = {
    0: {'label': '手机下载', 'url': '/downlist1'},
    1: {'label': '中文字幕', 'url': '/movielist1'},
    2: {'label': '亚洲五码', 'url': '/movielist2'},
    3: {'label': '欧美精品', 'url': '/movielist3'},
    4: {'label': '动漫其他', 'url': '/movielist4'}
}


# html转换
def html_to_soup(url):
    r = requests.get(url)
    html = r.content
    return BeautifulSoup(html, 'xml')


# url构造器 获得爬取链接
def url_constructor(page_index, _type):
    # https://www.971cf.com/htm/movielist1/2.htm
    aim_url = base_url + mid + type_dict[_type]['url'] + '/' + str(page_index) + suffix
    return aim_url


# 获取链接页面所包含的详情页链接并封装到List
def details_url_list_getter(url):
    global real_max_page_num  # 声明全局变量

    soup = html_to_soup(url)

    details_url_list = []
    movie_block_list = soup.findAll('li')
    for movie_block in movie_block_list:
        details_url_list.append(base_url + movie_block.find('a')['href'])

    # 获取页码信息,修正爬取页数
    page_info_block = soup.find('div', attrs={'class': 'pagination'})
    page_info_number = page_info_block.findAll('a')
    real_max_page_num = int(page_info_number[len(page_info_number) - 1]['href'].replace(suffix, ''))

    return details_url_list


# 详情页信息提取
def details_info_getter(details_url):
    random_sleep()  # 随机防ban
    browser.get(details_url)
    page_source = browser.page_source
    # 有时候网页会夏姬八跳
    if '发送任意内容邮件给' not in page_source:
        return None
    soup = BeautifulSoup(page_source, 'xml')

    result_dic = dict()  # 返回爬取结果字典

    film_info_block = soup.find('div', attrs={'class': 'film_info clearfix'})
    film_title = film_info_block.find('dd', attrs={'class': 'film_title'}).get_text()

    # 电影名
    result_dic['film_title'] = film_title

    download_info_block = soup.find('ul', id='downUL')
    # 下载地址
    download_url = download_info_block.find('input', attrs={'name': 'CopyAddr1'})['value']

    result_dic['download_url'] = download_url
    return result_dic


# 爬取逻辑
def data_spider(_choice, total_page=100):
    page_num = 1
    while page_num <= total_page:
        page_info = []
        url_list = details_url_list_getter(url_constructor(page_num, _choice))
        for url in url_list:
            temp = details_info_getter(url)
            if temp is not None:
                page_info.append(temp)
        data_output_xls(page_info)
        page_num += 1
        random_sleep()  # 随机防ban


# 输出数据到excel
def data_output_xls(data_list):
    global row
    if not os.path.exists(file_name):
        wb = Workbook()
    else:
        wb = load_workbook(file_name)
    title = "Hentai"
    try:
        work_sheet = wb[title]
    except KeyError:
        work_sheet = wb.create_sheet(title=title)
    # 标题行
    _ = work_sheet.cell(column=1, row=1, value="%s" % '电影名称')
    _ = work_sheet.cell(column=2, row=1, value="%s" % '链接地址')

    # 数据行
    for it in data_list:
        _ = work_sheet.cell(column=1, row=row, value="%s" % it['film_title'])  # 电影名称
        _ = work_sheet.cell(column=2, row=row, value="%s" % it['download_url'])  # 电影名称
        row += 1
    wb.save(filename=file_name)
    print('数据输出完成....')


# 随机休眠，防ban
def random_sleep():
    sleep_second = random.randint(0, 2)
    print('随机休眠' + str(sleep_second) + '秒...')
    time.sleep(sleep_second)


# Main method
if __name__ == '__main__':
    for i in range(len(type_dict)):
        print(i, type_dict[i]['label'], sep='.')
    choice = input('选择类型(任意字符表示0)：')
    try:  # 避免老铁夏姬八打
        if 0 < int(choice) < len(type_dict):
            choice = int(choice)
        else:
            choice = 0
    except ValueError:
        choice = 0
    data_spider(choice)
    # 任务完成后退出
    browser.quit()
